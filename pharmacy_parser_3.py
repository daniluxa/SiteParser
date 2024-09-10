import string
import requests
from bs4 import BeautifulSoup as bs
import requests
import openpyxl as ox
import time
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

pharm_name = []
pharm_price = []

url = 'https://www.acmespb.ru/pharma/oz17'
availability_of_the_next_page = True

catalog_name = ['a', 'b', 'v', 'g', 'd', 'e', 'zh', 'z', 'i', 'y', 'k', 'l', 'm', 'n', 'o', 'p', 'r', 
                's', 't', 'u', 'f', 'kh', 'ts', 'ch', 'sh', 'sch', 'ye', 'yu', 'ya', '09', 'az']

catalog_name_cnt = 0
page_cnt = 2
last_page = -1

buttons = []

new_url = f"{url}?alpha_code={catalog_name[catalog_name_cnt]}"

service = Service(executable_path='./geckodriver.exe')
options = webdriver.FirefoxOptions()
driver = webdriver.Firefox(options=options)

driver.get(new_url)

while(availability_of_the_next_page):

    # Получаем HTML-код страницы
    page = driver.page_source

    soup = bs(page, "html.parser")

    if(last_page == -1):
        count_of_find_pages = soup.find_all("span", class_="page")  # извлекаем кол-во страниц 
        try:
            last_page = int(count_of_find_pages[-1].text)
        except:
            last_page = 0

    tmp_pharm_name = soup.find_all("div", class_="cell name")  # извлекаем название 
    tmp_pharm_price = soup.find_all("div", class_="cell pricefull")  # извлекаем цену

    for data in tmp_pharm_name:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        pharm_name.append(tmp)
         
    for data in tmp_pharm_price:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        pharm_price.append(tmp)

    #button = driver.find_elements(By.CLASS_NAME, "page")

    #page_cnt = 2
    
    try:
        button = driver.find_element(By.XPATH, F"/html/body/div[2]/div/div[1]/div[2]/div[3]/div/div[5]/p[2]/span[{page_cnt}]")
    except:
        button = None

    if page_cnt < last_page:
        button = driver.find_element(By.XPATH, F"/html/body/div[2]/div/div[1]/div[2]/div[3]/div/div[5]/p[2]/span[{page_cnt}]")
        page_cnt += 1
        button.click()
    elif catalog_name_cnt < len(catalog_name) - 1:
        catalog_name_cnt += 1
        page_cnt = 2
        last_page = -1
        new_url = f"{url}?alpha_code={catalog_name[catalog_name_cnt]}"
        driver.get(new_url)
    else:
        availability_of_the_next_page = False

driver.quit()

wb = ox.load_workbook('parsing_acm.xlsx')
ws = wb.worksheets[0]
#ws.delete_cols(9)
ws.cell(row=1, column=9).value = "Название" 
for i, statN in enumerate(pharm_name): 
    ws.cell(row=i+2, column=9).value = statN
ws.cell(row=1, column=10).value = "Цена" 
for i, statN in enumerate(pharm_price): 
    ws.cell(row=i+2, column=10).value = statN 
wb.save('parsing_acm.xlsx')
