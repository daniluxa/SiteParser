import requests
from bs4 import BeautifulSoup as bs
import requests
import openpyxl as ox
import time
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options


pharm_name = []
pharm_price = []

url = 'https://ozerki.ru/sankt-peterburg/catalog'
availability_of_the_next_page = True

catalog_name = ['lekarstva-ot-prostudy-i-grippa', 'preparaty-dlya-pishchevaritelnogo-trakta','preparaty-dlya-serdechno-sosudistoj-sistemy', 
                'preparaty-pri-boleznyah-nervnoj-sistemy','planirovanie_semi', 'fitopreparaty','mama_i_malysh','zdorovoe_pitanie','vitaminy-i-obmen-veshchestv',
                'preparaty-dlya-lecheniya-zabolevanij-kozhi', 'sredstva-ot-boli-vospaleniya-temperatury','ginekologicheskie-preparaty', 
                'preparaty-dlya-kostno-myshechnoj-sistemy','preparaty-dlya-mochepolovoj-sistemy','preparaty-dlya-zreniya-i-sluha','preparaty-pri-allergii',
                'preparaty-dlya-lecheniya-ehndokrinnoj-sistemy','preparaty-pri-zabolevaniyah-krovi', 'preparaty-dlya-lecheniya-infekcij',
                'onkologiya-i-immunologiya','sredstva-ot-varikoza','preparaty-pri-zabolevaniyah-legkih','preparaty-dlya-borby-s-vrednymi-privychkami','preparaty-ot-parazitov',
                'bazovaya-fitoterapiya','gomeopaticheskie-sredstva','vakciny-i-syvorotki']

catalog_name_cnt = 0 
page_num = 99
last_page = -1

new_url = f"{url}/{catalog_name[catalog_name_cnt]}/?page={page_num}"

driver = webdriver.Chrome()


while(availability_of_the_next_page):
    driver.get(new_url)

    # Получаем HTML-код страницы
    page = driver.page_source

    soup = bs(page, "html.parser")

    if(last_page == -1):
        count_of_find_pages = soup.find_all("a", class_="AppRouterLink_link__uudGk sc-b9a2ebac-0 cpVAFQ")  # извлекаем кол-во страниц 
        last_page = int(count_of_find_pages[-1].text)

    tmp_pharm_name = soup.find_all("a", class_="AppRouterLink_link__uudGk sc-128b053f-1 pvICS product-name")  # извлекаем название 
    tmp_pharm_price = soup.find_all("div", class_="product-price__base-price")  # извлекаем цену

    for data in tmp_pharm_name:
         pharm_name.append(data.text)
         
    for data in tmp_pharm_price:
        amper_index = str(data.text).find('&')
        pharm_price.append(data.text[:amper_index])
    ##status = soup.find("span", class_="b-pagination-vuetify-imitation__item b-pagination-vuetify-imitation__item_next b-pagination-vuetify-imitation__item_disabled")

    if page_num < last_page:
        page_num += 1
        new_url = f"{url}/{catalog_name[catalog_name_cnt]}/?page={page_num}"
    elif catalog_name_cnt <= len(catalog_name):
        catalog_name_cnt += 1
    else:
        availability_of_the_next_page = False

driver.quit()

wb = ox.load_workbook('parsing_ozerki.xlsx')
ws = wb.worksheets[0]
ws.cell(row=1, column=9).value = "Название" 
for i, statN in enumerate(pharm_name): 
    ws.cell(row=i+2, column=9).value = statN 
ws.cell(row=1, column=10).value = "Цена" 
for i, statN in enumerate(pharm_price): 
    ws.cell(row=i+2, column=10).value = statN 
wb.save('parsing_ozerki.xlsx')