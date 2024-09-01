import string
import requests
from bs4 import BeautifulSoup as bs
import requests
import openpyxl as ox
import time
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

pharm_name = []
pharm_price = []

url = 'https://lekopttorg.ru'
availability_of_the_next_page = True

#catalog_name = ['lekarstva_i_profilakticheskie_sredstva', 'krasota_ukhod_i_gigiena','meditsinskie_tovary', 'meditsinskaya_tekhnika','planirovanie_semi', 'fitopreparaty','mama_i_malysh','zdorovoe_pitanie'
#                ,'optika_i_kontaktnaya_korrektsiya','redkie_preparaty', 'upakovka_new','lechebnoe_pitanie', 'prostyni_i_pelenki_vpityvayushchie']
#catalog_name_cnt = 0
page_num = 1

new_url = f"{url}/catalog/?by=1000%2Fpage%3D5&PAGEN_3={page_num}"

driver = webdriver.Chrome()

while(availability_of_the_next_page):
    driver.get(new_url)

    # Получаем HTML-код страницы
    page = driver.page_source

    soup = bs(page, "html.parser")

    tmp_pharm_name = soup.find_all("a", class_="product__title mb-8 title title_block")  # извлекаем название 
    tmp_pharm_price = soup.find_all("span", class_="price__regular")  # извлекаем цену

    for data in tmp_pharm_name:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        pharm_name.append(tmp)
         
    for data in tmp_pharm_price:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        pharm_price.append(tmp)

    ##status = soup.find("span", class_="b-pagination-vuetify-imitation__item b-pagination-vuetify-imitation__item_next b-pagination-vuetify-imitation__item_disabled")
    status = soup.find("div", class_="arrow__right")
    availability_of_the_next_page = False
    if status != None:
        page_num += 1
        new_url = f"{url}/catalog/?by=1000%2Fpage%3D5&PAGEN_3={page_num}"
    else:
        availability_of_the_next_page = False

driver.quit()

wb = ox.load_workbook('parsing.xlsx')
ws = wb.worksheets[0]
ws.cell(row=1, column=9).value = "Название" 
for i, statN in enumerate(pharm_name): 
    ws.cell(row=i+2, column=9).value = statN
ws.cell(row=1, column=10).value = "Цена" 
for i, statN in enumerate(pharm_price): 
    ws.cell(row=i+2, column=10).value = statN 
wb.save('parsing.xlsx')
