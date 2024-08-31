import requests
from bs4 import BeautifulSoup as bs
import requests
import openpyxl as ox
import time
from selenium import webdriver

Doctors_Name = []
Doctors_Spec = []
Doctors_Categoria  = []
Doctors_Experience = []
Doctors_Clinic = []
Doctors_Address = []

url = 'https://prodoctorov.ru/spb/vrach/'
new_page_url = '?page='
availability_of_the_next_page = True
page_num = 1

new_url = url
driver = webdriver.Chrome()

while(availability_of_the_next_page):
    driver.get(new_url)

    # Получаем HTML-код страницы
    page = driver.page_source

    soup = bs(page, "html.parser")

    tmp_Doctors_Name       = soup.find_all("span", class_="b-doctor-card__name-surname")  # извлекаем ФИО
    tmp_Doctors_Spec       = soup.find_all("div", class_="b-doctor-card__spec")  # извлекаем Специфику
    tmp_Doctors_Categoria  = soup.find_all("div", class_="b-doctor-card__category")  # извлекаем Опыт
    tmp_Doctors_Experience = soup.find_all("div", class_="b-doctor-card__experience-years")  # извлекаем Стаж
    tmp_Doctors_Clinic     = soup.find_all("span", class_="b-select__trigger-main-text")  # извлекаем Больницу
    tmp_Doctors_Address    = soup.find_all("span", class_="b-select__trigger-adit-text")  # извлекаем адрес больницы

    for data in tmp_Doctors_Name:
        Doctors_Name.append(data.text)

    for data in tmp_Doctors_Spec:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        Doctors_Spec.append(tmp)

    for data in tmp_Doctors_Categoria:
        Doctors_Categoria.append(data.text)

    for data in tmp_Doctors_Experience:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        tmp = tmp[5::]
        Doctors_Experience.append(tmp)

    for data in tmp_Doctors_Clinic:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        Doctors_Clinic.append(tmp)
    for data in tmp_Doctors_Address:
        data = data.text.replace('\n', '')
        tmp = ' '.join(data.split())
        Doctors_Address.append(tmp)

    status = soup.find("span", class_="b-pagination-vuetify-imitation__item b-pagination-vuetify-imitation__item_next b-pagination-vuetify-imitation__item_disabled")
    if status == None:
        page_num += 1
        new_url = url + new_page_url +str(page_num)
    else:
        availability_of_the_next_page = False

driver.quit()

wb=ox.Workbook()
ws = wb.worksheets[0]
ws.cell(row=1, column=1).value = "ФИО" 
ws.cell(row=1, column=2).value = "Специфика" 
ws.cell(row=1, column=3).value = "Категория" 
ws.cell(row=1, column=4).value = "Стаж"
ws.cell(row=1, column=5).value = "Название клиники"
ws.cell(row=1, column=6).value = "Адрес клиники"
for i, statN in enumerate(Doctors_Name): 
    ws.cell(row=i+2, column=1).value = statN 
for i, statN in enumerate(Doctors_Spec): 
    ws.cell(row=i+2, column=2).value = statN 
for i, statN in enumerate(Doctors_Categoria): 
    ws.cell(row=i+2, column=3).value = statN
for i, statN in enumerate(Doctors_Experience): 
    ws.cell(row=i+2, column=4).value = statN
for i, statN in enumerate(Doctors_Clinic): 
    ws.cell(row=i+2, column=5).value = statN
for i, statN in enumerate(Doctors_Address): 
    ws.cell(row=i+2, column=6).value = statN
wb.save('doctors.xlsx')