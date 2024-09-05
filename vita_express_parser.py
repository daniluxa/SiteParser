import requests
import time
from bs4 import BeautifulSoup as bs
import requests
import openpyxl as ox
import time
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By

pharm_name = []
pharm_price = []

url = 'https://ozerki.ru/sankt-peterburg/catalog'
availability_of_the_next_page = True

url = 'https://vitaexpress.ru/catalog/lekarstva-i-bady/lekarstva-ot-prostudy/'

catalog_name = ['lekarstva-ot-prostudy-i-grippa', 'preparaty-dlya-pishchevaritelnogo-trakta','preparaty-dlya-serdechno-sosudistoj-sistemy', 
                'preparaty-pri-boleznyah-nervnoj-sistemy','planirovanie_semi', 'fitopreparaty','mama_i_malysh','zdorovoe_pitanie','vitaminy-i-obmen-veshchestv',
                'preparaty-dlya-lecheniya-zabolevanij-kozhi', 'sredstva-ot-boli-vospaleniya-temperatury','ginekologicheskie-preparaty', 
                'preparaty-dlya-kostno-myshechnoj-sistemy','preparaty-dlya-mochepolovoj-sistemy','preparaty-dlya-zreniya-i-sluha','preparaty-pri-allergii',
                'preparaty-dlya-lecheniya-ehndokrinnoj-sistemy','preparaty-pri-zabolevaniyah-krovi', 'preparaty-dlya-lecheniya-infekcij',
                'onkologiya-i-immunologiya','sredstva-ot-varikoza','preparaty-pri-zabolevaniyah-legkih','preparaty-dlya-borby-s-vrednymi-privychkami','preparaty-ot-parazitov',
                'bazovaya-fitoterapiya','gomeopaticheskie-sredstva','vakciny-i-syvorotki']

catalog_name_cnt = 0 
page_num = 1
last_page = -1

new_url = f"{url}/{catalog_name[catalog_name_cnt]}/?page={page_num}"

service = Service(executable_path='./chromedriver.exe')
options = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=options)

#driver = webdriver.Chrome()

#<button id="newBtnPager" class="mt-30 btn-pager">Показать ещё</button>

while(availability_of_the_next_page):
    #driver.get(new_url)
    driver.get(url)

    last_height = driver.execute_script("return document.body.scrollHeight") 
    while True: 
        # Прокрутка вниз 
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);") 
        # Пауза, пока загрузится страница. 
        time.sleep(2)
        # Вычисляем новую высоту прокрутки и сравниваем с последней высотой прокрутки. 
        new_height = driver.execute_script("return document.body.scrollHeight") 
        if new_height == last_height: 
            button = driver.find_element(By.ID, "newBtnPager")
            button.click()
            if(button == None):
                print("Прокрутка завершена") 
                break 
        last_height = new_height
        print("Появился новый контент, прокручиваем дальше")

    # Получаем HTML-код страницы
    page = driver.page_source

    soup = bs(page, "html.parser")

    #tmp_pharm_name = soup.find_all("div", class_='horizontalCard__info')  # извлекаем название 

    tmp_pharm_name = soup.find_all("div", class_="relative mb-8")

    #tmp_pharm_price = soup.find_all("div", class_="product-price__base-price")  # извлекаем цену

    for data in tmp_pharm_name:
         pharm_name.append(data.text)
         
    # for data in tmp_pharm_price:
    #     amper_index = str(data.text).find('&')
    #     pharm_price.append(data.text[:amper_index])
    ##status = soup.find("span", class_="b-pagination-vuetify-imitation__item b-pagination-vuetify-imitation__item_next b-pagination-vuetify-imitation__item_disabled")

    # if page_num < last_page:
    #     page_num += 1
    #     new_url = f"{url}/{catalog_name[catalog_name_cnt]}/?page={page_num}"
    # elif catalog_name_cnt <= len(catalog_name):
    #     catalog_name_cnt += 1
    #     page_num = 1
    #     last_page = -1
    #     new_url = f"{url}/{catalog_name[catalog_name_cnt]}/?page={page_num}"
    # else:
    #     availability_of_the_next_page = False

    availability_of_the_next_page = False

driver.quit()

wb = ox.load_workbook('parsing_ozerki.xlsx')
ws = wb.worksheets[0]
ws.cell(row=1, column=9).value = "Название" 
for i, statN in enumerate(pharm_name): 
    ws.cell(row=i+2, column=9).value = statN 
ws.cell(row=1, column=10).value = "Цена" 
for i, statN in enumerate(pharm_price): 
    ws.cell(row=i+2, column=10).value = statN 
wb.save('parsing_ozerki.xlsx')